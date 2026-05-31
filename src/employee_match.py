"""
员工批量导出 — 读取 Excel 员工表，匹配微信联系人，批量导出聊天记录。
"""
import os
import sys
from collections import defaultdict

from chat_export import export_chat
from chat_list import scan_chats
from engine.utils import load_contacts

# 列名映射配置: 逻辑字段 → [表头匹配模式列表]，按优先级排列
# 新增 Excel 格式时只需在此添加模式即可
_COLUMN_MAP = {
    'name':   ['姓名', '员工姓名', '用户名', '名字', '名称'],
    'dept':   ['部门', '所属部门', '组织', '机构', '部门名称'],
    'phone':  ['移动电话', '手机', '电话', '手机号', '联系电话', '联系方式'],
    'status': ['禁用', '状态', '启用', '在职状态', '账号状态'],
    'region': ['大区', '区域', '地区', '片区', '所属大区'],
}


def _detect_columns(headers):
    """根据 _COLUMN_MAP 配置检测 Excel 列索引。
    对每个逻辑字段，按模式列表顺序匹配，首个命中即返回。
    name 字段若未匹配到则默认返回列 0；其余字段返回 None。
    """
    cols = {}
    for field, patterns in _COLUMN_MAP.items():
        for pat in patterns:
            for i, h in enumerate(headers):
                if pat in h:
                    cols[field] = i
                    break
            if field in cols:
                break
        else:
            cols[field] = None
    return cols


def load_employees(excel_path):
    """从 Excel 加载员工信息。"""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要 openpyxl 库: pip install openpyxl")
        return []

    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel 文件不存在: {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    cols = _detect_columns(headers)

    employees = []
    name_col = cols.get('name')
    if name_col is None:
        print("[ERROR] 无法识别姓名列，请检查 Excel 表头")
        return []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_col] if name_col < len(row) else None
        if name and str(name).strip():
            employees.append({
                'name': str(name).strip(),
                'dept': str(row[cols['dept']]).strip() if cols['dept'] is not None and cols['dept'] < len(row) and row[cols['dept']] else '',
                'phone': str(row[cols['phone']]).strip() if cols['phone'] is not None and cols['phone'] < len(row) and row[cols['phone']] else '',
                'status': str(row[cols['status']]).strip() if cols['status'] is not None and cols['status'] < len(row) and row[cols['status']] else '',
                'region': str(row[cols['region']]).strip() if cols['region'] is not None and cols['region'] < len(row) and row[cols['region']] else '',
            })
    return employees


def load_all_contacts(decrypted_dir):
    """加载所有联系人信息用于匹配。"""
    contacts = []
    id_to_name, name_to_id, usernames = load_contacts(decrypted_dir)

    contact_db = os.path.join(decrypted_dir, "contact", "contact.db")
    if not os.path.exists(contact_db):
        return contacts

    import sqlite3
    conn = sqlite3.connect(contact_db)
    try:
        for r in conn.execute(
            "SELECT id, username, remark, nick_name, alias, description FROM contact"
        ):
            cid, uname, remark, nick, alias, desc = r
            remark = (remark or "").strip()
            nick = (nick or "").strip()
            alias = (alias or "").strip()
            desc = (desc or "").strip()
            display = remark or nick or alias or uname
            contacts.append({
                'id': cid,
                'username': (uname or "").strip(),
                'remark': remark,
                'nick_name': nick,
                'alias': alias,
                'description': desc,
                'display_name': display,
            })
    finally:
        conn.close()
    return contacts


def score_match(emp_name, c):
    """评分员工与联系人的匹配度。"""
    en = emp_name

    if c['remark'] == en:
        return (100, f"备注完全匹配: {c['remark']}")
    if c['alias'] == en:
        return (90, f"别名完全匹配: {c['alias']}")
    if en in c['remark'] and len(en) >= 2:
        return (85, f"备注包含姓名: {c['remark']}")
    if c['nick_name'] == en:
        return (80, f"昵称完全匹配: {c['nick_name']}")
    if en in c['nick_name'] and len(en) >= 2:
        return (70, f"昵称包含姓名: {c['nick_name']}")
    if c['description'] == en:
        return (60, f"描述完全匹配: {c['description']}")
    if c['nick_name'] and len(c['nick_name']) >= 2 and c['nick_name'] in en:
        return (50, f"姓名包含昵称: {c['nick_name']}")
    if c['remark'] and len(c['remark']) >= 2 and c['remark'] in en:
        return (45, f"姓名包含备注: {c['remark']}")
    if en in c['description'] and len(en) >= 2:
        return (40, f"描述包含姓名: {c['description']}")
    if c['description'] and len(c['description']) >= 2 and c['description'] in en:
        return (30, f"姓名包含描述: {c['description']}")

    return (0, "")


def match_employees(employees, contacts, min_score=30):
    """匹配员工到微信联系人。"""
    matches = {}
    unmatched = []
    contact_used = defaultdict(list)

    for emp in employees:
        en = emp['name']
        best_score = 0
        best_contact = None
        best_reason = ""

        for c in contacts:
            score, reason = score_match(en, c)
            if score > best_score:
                best_score = score
                best_contact = c
                best_reason = reason

        if best_score >= min_score and best_contact:
            matches[en] = {
                'employee': emp,
                'contact': best_contact,
                'score': best_score,
                'reason': best_reason,
            }
            contact_used[best_contact['id']].append(en)
        else:
            unmatched.append(en)

    for en, m in matches.items():
        cid = m['contact']['id']
        m['is_ambiguous'] = len(contact_used[cid]) > 1

    return matches, unmatched


def run_employee_export(decrypted_dir, excel_path, out_dir, start_ts=None, end_ts=None,
                        keyword=None, min_score=30, name_filter=None, list_only=False,
                        print_fn=None, progress_fn=None):
    """主入口：员工批量导出。"""
    if print_fn is None:
        print_fn = print
    if progress_fn is None:
        progress_fn = lambda pct, msg: None

    # Load employees
    progress_fn(5, "加载员工信息...")
    employees = load_employees(excel_path)
    if not employees:
        return
    print_fn(f"共 {len(employees)} 名员工")

    # Load contacts
    progress_fn(15, "加载微信联系人...")
    contacts = load_all_contacts(decrypted_dir)
    print_fn(f"共 {len(contacts)} 个微信联系人")

    # Match
    progress_fn(25, "匹配员工与联系人...")
    matches, unmatched = match_employees(employees, contacts, min_score)
    print_fn(f"匹配成功: {len(matches)} 人, 未匹配: {len(unmatched)} 人")

    if unmatched:
        print_fn(f"\n未匹配人员 ({len(unmatched)}):")
        for name in unmatched:
            print_fn(f"  - {name}")

    if list_only:
        return matches, unmatched

    # Filter by name
    if name_filter:
        names = set(n.strip() for n in name_filter.split(","))
        matches = {n: m for n, m in matches.items() if n in names}
        if not matches:
            print_fn("指定的姓名未匹配到任何微信联系人")
            return matches, unmatched

    # Scan chats to find matching usernames
    progress_fn(35, "扫描聊天列表...")
    chats, _, _ = scan_chats(decrypted_dir)
    chat_by_username = {c["username"]: c for c in chats}

    # Export
    os.makedirs(out_dir, exist_ok=True)
    total = len(matches)
    exported = 0
    no_msgs = 0
    total_msgs = 0

    for i, (name, m) in enumerate(sorted(matches.items())):
        pct = 40 + int((i + 1) / total * 55)
        progress_fn(pct, f"导出: {name} ({i+1}/{total})")

        uname = m['contact']['username']
        chat = chat_by_username.get(uname)
        if not chat:
            print_fn(f"  {name}: 未找到聊天记录")
            no_msgs += 1
            continue

        count, path = export_chat(chat, out_dir, start_ts, end_ts, keyword,
                                  print_fn=None)
        if count > 0:
            # Add employee header to file
            emp = m['employee']
            header = (
                f"员工: {emp['name']}\n"
                f"部门: {emp['dept']}\n"
                f"微信联系人: {m['contact']['display_name']}\n"
                f"微信号: {m['contact']['username']}\n"
                f"匹配方式: {m['reason']} (置信度: {m['score']})\n"
            )
            if m.get('is_ambiguous'):
                header += "⚠ 注意: 此匹配可能存在歧义，请核实\n"

            # Prepend header to file (write-temp + rename for atomicity)
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()
            tmp_path = path + '.tmp'
            with open(tmp_path, 'w', encoding='utf-8') as f:
                f.write(header + '=' * 70 + '\n\n' + content.split('=' * 70 + '\n\n', 1)[-1])
            os.replace(tmp_path, path)

            print_fn(f"  {name}: {count} 条消息 -> {os.path.basename(path)}")
            total_msgs += count
            exported += 1
        else:
            print_fn(f"  {name}: 无消息")
            no_msgs += 1

    progress_fn(98, f"导出完成: {exported}人/{total_msgs}条消息")
    print_fn(f"\n导出完成! {exported}人/{total_msgs}条消息, 无消息:{no_msgs}, 未匹配:{len(unmatched)}")

    return matches, unmatched
